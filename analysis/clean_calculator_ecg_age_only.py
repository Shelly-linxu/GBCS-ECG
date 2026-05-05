from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def clean_calculator(input_xlsx: Path, output_xlsx: Path) -> None:
    """Remove residual ECG-age acceleration content from the supplementary calculator."""
    wb = load_workbook(input_xlsx, data_only=False)

    readme = wb["README"]
    for row in readme.iter_rows():
        for cell in row:
            cell.value = None
    readme_rows = [
        ("ECG Age Calculator", ""),
        ("", ""),
        (
            "Purpose",
            "Use the Input_Data sheet to enter one participant per row. The Scores sheet calculates exact LightGBM ECG age from the exported tree equations.",
        ),
        (
            "Rows supported",
            "Rows 6-55 are pre-formulated. Copy formulas down in Input_Imputed, TreeScores, and Scores if more rows are required.",
        ),
        (
            "Missing ECG values",
            "Leave an ECG feature blank if missing. The workbook imputes blank ECG feature cells to model-development cohort medians.",
        ),
        ("Do not edit", "TreeScores, Input_Imputed, Variables, and Model_Metadata contain model logic and constants."),
        (
            "Required non-ECG input",
            "case_id is required for a row to calculate. agec and sex are retained as participant descriptors but are not used by the ECG-age model.",
        ),
        ("ECG age", "Exact final LightGBM model from ecg_age_best_model_full_fit."),
        (
            "Caution",
            "This workbook reproduces the structured-ECG measurement model used in the manuscript and is intended for research use and external validation.",
        ),
    ]
    for row_index, values in enumerate(readme_rows, 1):
        readme.cell(row_index, 1).value = values[0]
        readme.cell(row_index, 2).value = values[1]

    scores = wb["Scores"]
    scores.delete_cols(5, 5)
    headers = [
        "case_id",
        "agec",
        "sex",
        "ecg_age",
        "n_ecg_features_expected",
        "n_ecg_features_provided",
        "n_ecg_features_missing",
    ]
    for col_index, header in enumerate(headers, 1):
        scores.cell(5, col_index).value = header
    for row_index in range(6, 56):
        scores.cell(row_index, 1).value = f'=IF(Input_Data!$A{row_index}="","",Input_Data!A{row_index})'
        scores.cell(row_index, 2).value = f'=IF(Input_Data!$A{row_index}="","",Input_Data!B{row_index})'
        scores.cell(row_index, 3).value = f'=IF(Input_Data!$A{row_index}="","",Input_Data!C{row_index})'
        scores.cell(row_index, 4).value = f'=IF(Input_Data!$A{row_index}="","",SUM(TreeScores!B{row_index}:SG{row_index}))'
        scores.cell(row_index, 5).value = f'=IF(Input_Data!$A{row_index}="","",174)'
        scores.cell(row_index, 6).value = f'=IF(Input_Data!$A{row_index}="","",COUNT(Input_Data!D{row_index}:FU{row_index}))'
        scores.cell(row_index, 7).value = f'=IF(Input_Data!$A{row_index}="","",E{row_index}-F{row_index})'
    for row_index in range(1, 56):
        for col_index in range(8, 13):
            scores.cell(row_index, col_index).value = None

    metadata = wb["Model_Metadata"]
    for row in metadata.iter_rows():
        for cell in row:
            cell.value = None
    metadata_rows = [
        ("Model metadata", ""),
        ("", ""),
        ("Key", "Value"),
        ("ECG age model", "LightGBMRegressor"),
        ("Number of ECG variables", 174),
        ("Number of LightGBM trees", 500),
        ("Tree leaves", "up to 31"),
        ("Missing value handling", "Blank ECG feature cells are imputed to model-development cohort medians."),
        ("Model source", "ecg_age_best_model_full_fit.joblib"),
        ("Feature list source", "ecg_age_feature_list.csv"),
        ("Reference cohort N", 17090),
        ("Formula: ECG age", "SUM of 500 LightGBM tree leaf values after median imputation"),
    ]
    for row_index, values in enumerate(metadata_rows, 1):
        metadata.cell(row_index, 1).value = values[0]
        metadata.cell(row_index, 2).value = values[1]

    validation = wb["Validation_Example"]
    for row_index in range(7, 15):
        validation.cell(row_index, 1).value = None
        validation.cell(row_index, 2).value = None
    validation["B5"].value = "=Scores!D6"
    validation["B6"].value = "=B5-B4"

    wrap = Alignment(wrap_text=True, vertical="top")
    for sheet_name in ["README", "Model_Metadata", "Validation_Example"]:
        sheet = wb[sheet_name]
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 82
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = wrap

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    clean_calculator(
        root / "supplementary/calculator/ecg_age_score_calculator_with_acceleration.xlsx",
        root / "supplementary/calculator/ecg_age_score_calculator.xlsx",
    )
